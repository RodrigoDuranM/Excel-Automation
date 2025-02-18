import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from win32com.client import Dispatch

def order_based_on_header(df_or_sheet, header, filter_order):
    """
    Orders a DataFrame or sheet based on a given header.
    filter_order: "high" (descending) or "low" (ascending)
    """
    if isinstance(df_or_sheet, pd.DataFrame):
        df = df_or_sheet.copy()
    else:
        df = pd.DataFrame(df_or_sheet.values)
        df.columns = df.iloc[0]
        df = df[1:].reset_index(drop=True)
    
    ascending = True if filter_order.lower() == "low" else False
    return df.sort_values(by=header, ascending=ascending)

def list_high_to_low(df, header):
    """Returns a list of values from highest to lowest based on the specified header column."""
    return df.sort_values(by=header, ascending=False)[header].tolist()

def list_low_to_high(df, header):
    """Returns a list of values from lowest to highest based on the specified header column."""
    return df.sort_values(by=header, ascending=True)[header].tolist()

# Function to create workbook and sheet (with a sheet title)
def create_workbook_and_sheet(sheet_title):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title
    return workbook, sheet

# Function to load the Excel file into a DataFrame
def excel_to_panda(excel_file):
    df = pd.read_excel(excel_file)
    return df

# Function to convert a Pandas DataFrame to a sheet in the Excel file
def panda_to_excel(data, sheet):
    # Create a Pandas DataFrame from the provided data
    df = pd.DataFrame(data)
    for row in dataframe_to_rows(df, index=False, header=True):
        sheet.append(row)
    return sheet

# Function to automate styling of the sheet
def style_sheet(sheet, letter_color, back_color, border_thickness, header_alignment, data_alignment, bold_headers):
    color_map = {
        "white": "FFFFFF",
        "black": "000000",
        "red": "FF0000",
        "green": "00FF00",
        "blue": "0000FF"   
    }
    border_map = {
        "thin": "thin",
        "medium": "medium",
        "thick": "thick"
    }
    alignment_map = {
        "left": "left",
        "center": "center",
        "right": "right"
    }
    
    font_color = color_map.get(letter_color.lower(), "000000")  # Default to Black if not found
    fill_color = color_map.get(back_color.lower(), "FFFFFF")  # Default to White if not found
    border_size = border_map.get(border_thickness.lower(), "thin")  # Default to Thin if not found
    header_alignment = alignment_map.get(header_alignment.lower(), "left")  # Default to Left if not found
    data_alignment = alignment_map.get(data_alignment.lower(), "left")  # Default to Left if not found
    if bold_headers == 'yes':
        bold = True
    else:
        bold = False
    header_font = Font(bold, color=font_color)  
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")  
    thin_border = Border(
        left=Side(style=border_size), 
        right=Side(style=border_size), 
        top=Side(style=border_size), 
        bottom=Side(style=border_size)
    )

    # Style each cell in the header row
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal=header_alignment)

    # Apply borders and alignment to the data rows
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal=data_alignment)

    # Auto-adjust column widths
    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get the column letter (A, B, C, etc.)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = max_length + 2

    return sheet

