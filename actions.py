
import os
import re
import pandas as pd
from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from formulas import style_sheet # File containing basic excel functions

def number_sorting(file, header, order, app):
    if not file or not header or not order:
            return "Missing file or parameters!", 400
        
    try:
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(file_path)

        # Load the Excel file using openpyxl to preserve styles
        wb = load_workbook(file_path)
        ws = wb.active

        # Find the header column index
        df = pd.read_excel(file_path)
        if header not in df.columns:
            return f"Column '{header}' not found in the file!", 400
        
        header_index = df.columns.get_loc(header) + 1  # openpyxl is 1-based index
        ascending = True if order == 'asc' else False
        
        # Extract data rows (excluding headers)
        data = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2, values_only=False)]
        
        # Sort data based on the selected column
        data.sort(key=lambda x: x[header_index - 1] if isinstance(x[header_index - 1], (int, float)) else float('inf'), reverse=not ascending)
        
        # Write sorted data back while maintaining formatting
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, cell_value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)

        # Save the sorted file
        sorted_file_path = os.path.join(app.config['PROCESSED_FOLDER'], "sorted_" + file.filename)
        wb.save(sorted_file_path)

        return send_file(sorted_file_path, as_attachment=True, download_name=f"sorted_{file.filename}")
    
    except Exception as e:
        return f"Error processing the file: {str(e)}", 500
    
def word_sorting(file, header, order, app):
    if not file or not header or not order:
            return "Missing file or parameters!", 400

    try:
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(file_path)

        # Load the Excel file with openpyxl to preserve formatting
        wb = load_workbook(file_path)
        ws = wb.active

        df = pd.read_excel(file_path)
        if header not in df.columns:
            return f"Column '{header}' not found in the file!", 400

        ascending = True if order == 'asc' else False
        
        # Extract column data for sorting
        col_index = df.columns.get_loc(header) + 1  # Convert to 1-based index
        rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))
        sorted_rows = sorted(rows, key=lambda x: x[col_index - 1], reverse=not ascending)

        # Rewrite sorted data back while preserving styles
        for i, row_data in enumerate(sorted_rows, start=2):
            for j, value in enumerate(row_data, start=1):
                ws.cell(row=i, column=j, value=value)

        sorted_file_path = os.path.join(app.config['PROCESSED_FOLDER'], "sorted_" + file.filename)
        wb.save(sorted_file_path)

        return send_file(sorted_file_path, as_attachment=True, download_name=f"sorted_{file.filename}")

    except Exception as e:
        return f"Error processing the file: {str(e)}", 500
    
def file_filter(file, column, value, sort_order, app):
    if not file or not column or not value or not sort_order:
        return "Missing file or parameters!", 400

    try:
        # Save the uploaded file
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(file_path)

        # Load workbook and select the active sheet
        wb = load_workbook(file_path)
        ws = wb.active

        # Find the column index based on the header (column name)
        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            if column in row:
                header_row = row
                header_col_idx = row.index(column) + 1  # Convert to 1-based index
                break

        if not header_row:
            return f"Column '{column}' not found in the file!", 400

        # Convert input value to float
        try:
            value = float(value)
        except ValueError:
            return "Invalid number entered!", 400

        # Collect row indices to delete based on filter condition
        rows_to_delete = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            cell_value = row[header_col_idx - 1]  # Adjust for 0-based index

            # Ensure cell_value is a number and check condition
            if isinstance(cell_value, (int, float)):
                if (sort_order == "desc" and cell_value <= value) or (sort_order == "asc" and cell_value >= value):
                    rows_to_delete.append(row_idx)

        # Delete rows (in reverse to avoid shifting issues)
        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx)

        # Save the modified workbook
        filtered_file_path = os.path.join(app.config['PROCESSED_FOLDER'], "filtered_" + file.filename)
        wb.save(filtered_file_path)

        # Return file for download
        return send_file(filtered_file_path, as_attachment=True, download_name=f"filtered_{file.filename}")

    except Exception as e:
        return f"Error processing the file: {str(e)}", 500

    
def dup_remover(file, app):
    if not file:
            return "No file uploaded!", 400

    try:
        # Save uploaded file
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(file_path)

        # Load workbook and active sheet
        wb = load_workbook(file_path)
        ws = wb.active

        # Read all rows into a set to track unique rows
        seen_rows = set()
        rows_to_delete = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_tuple = tuple(row)  # Convert row to tuple to make it hashable
            if row_tuple in seen_rows:
                rows_to_delete.append(row_idx)  # Mark for deletion
            else:
                seen_rows.add(row_tuple)  # Add unique row

        # Delete duplicate rows (reverse order to prevent shifting issues)
        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx)

        # Save the modified workbook
        processed_file_path = os.path.join(app.config['PROCESSED_FOLDER'], "no_duplicates_" + file.filename)
        wb.save(processed_file_path)

        # Return file for download
        return send_file(processed_file_path, as_attachment=True, download_name=f"no_duplicates_{file.filename}")

    except Exception as e:
        return f"Error processing the file: {str(e)}", 500

def num_highlight(file, header, condition, key_value, app):
    if not file or not header or not condition or not key_value:
            return "Missing file or parameters!", 400
        
    try:
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(file_path)

        df = pd.read_excel(file_path)

        if header not in df.columns:
            return f"Column '{header}' not found in the file!", 400
        
        key_value = float(key_value)

        # Load the Excel file with openpyxl
        wb = load_workbook(file_path)
        ws = wb.active

        # Define the highlight color
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Get the column index for the specified header
        header_index = df.columns.get_loc(header) + 1  # openpyxl is 1-based index
        
        # Iterate through rows and apply highlighting
        for row in range(2, ws.max_row + 1):  # Start from 2 to skip header row
            cell_value = ws.cell(row=row, column=header_index).value
            if isinstance(cell_value, (int, float)):
                if (condition == 'desc' and cell_value > key_value) or (condition == 'asc' and cell_value < key_value):
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = highlight_fill

        # Save the highlighted file
        highlighted_file_path = os.path.join(app.config['PROCESSED_FOLDER'], "highlighted_" + file.filename)
        wb.save(highlighted_file_path)

        return send_file(highlighted_file_path, as_attachment=True, download_name=f"highlighted_{file.filename}")

    except Exception as e:
        return f"Error processing the file: {str(e)}", 500

def word_highlight(file, header, key_value, app):
    if not file or not header or not key_value:
            return "Missing file or parameters!", 400

    try:
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(file_path)

        # Load the Excel file with openpyxl
        wb = load_workbook(file_path)
        ws = wb.active

        # Check if the header exists
        df = pd.read_excel(file_path)
        if header not in df.columns:
            return f"Column '{header}' not found in the file!", 400
        
        # Define the highlight color
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Get the column index for the specified header
        header_index = df.columns.get_loc(header) + 1  # openpyxl is 1-based index
        
        # Iterate through rows and apply highlighting
        for row in range(2, ws.max_row + 1):  # Start from 2 to skip header row
            cell_value = ws.cell(row=row, column=header_index).value
            if isinstance(cell_value, str) and key_value.lower() in cell_value.lower():
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = highlight_fill

        # Save the highlighted file
        highlighted_file_path = os.path.join(app.config['PROCESSED_FOLDER'], "highlighted_" + file.filename)
        wb.save(highlighted_file_path)

        return send_file(highlighted_file_path, as_attachment=True, download_name=f"highlighted_{file.filename}")

    except Exception as e:
        return f"Error processing the file: {str(e)}", 500

def styling(file, font_color, header_background, bold_headers, border_thickness, header_alignment, data_alignment, app):
    if not file:
            return "No file uploaded", 400

    try:
        # Save the file temporarily
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(file_path)

        # Load the Excel file with openpyxl
        wb = load_workbook(file_path)
        ws = wb.active

        # Call the styling function
        styled_sheet = style_sheet(ws, font_color, header_background, border_thickness, header_alignment, data_alignment, bold_headers)
        ws = styled_sheet

        # Save the styled file
        styled_file_path = os.path.join(app.config['PROCESSED_FOLDER'], "styled_" + file.filename)
        wb.save(styled_file_path)

        return send_file(styled_file_path, as_attachment=True, download_name=f"styled_{file.filename}")

    except Exception as e:
        return f"Error processing the file: {str(e)}", 500

